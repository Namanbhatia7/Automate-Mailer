import os
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

os.chdir()         #Enter path of files in chdir()


workbook = openpyxl.load_workbook('sample.xlsx')  # open excel file

sheet = workbook.active
email = []            
candidate_name = []


for i in range(2,16):              #run loop 
    email.append(sheet.cell(row = i, column= 3).value)    # Take values from email coulumn. Change according to you .xlsx file.
    

for i in range(2,16):
     candidate_name.append(sheet.cell(row = i, column= 2).value)    #  values from Name column


def sendmail(candidate_name,email):

    
    sender_email= ''           # Enter your Email
    reciever_email= email

    msg = MIMEText('This is sample Message')         # Enter your message

    msg['From'] = sender_email
	 
	 
    msg['To'] = reciever_email
	 
	
    msg['Subject'] = ("This is subject field")           # Enter the subject of mail

    print('server')
    s = smtplib.SMTP('smtp-mail.outlook.com', 587)       # If sender email id is outlook.com, Check for smtp for other email clients

    print('server') 
    # start TLS for security
    s.starttls()
    
    password="python@123"                            #Enter your password
    # Authentication
    s.login(sender_email,password)
	 
	 
    # sending the mail
    s.sendmail(sender_email, reciever_email, msg.as_string())
    
	 
    # terminating the session
    s.quit()

if __name__ == "__main__":	
    for i in email:
      sendmail(candidate_name[i],email[i])
    
